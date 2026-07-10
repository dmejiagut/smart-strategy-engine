"""Generación e importación de la plantilla Excel de compras/ventas.

Diseño:
- Hoja 'Movimientos' (donde se llena): la Estrategia y la Operación se eligen de
  listas desplegables, no se escriben a mano. La lista de estrategias muestra el
  detalle que distingue a cada una (frecuencia, fechas, entrada/salida) para que
  dos estrategias del mismo ticker no se confundan, y lleva un id oculto.
- Hoja 'Copy Trading' (aparte): como Copy es una canasta por acción, tiene su
  propia hoja — eliges la Cartera y la Acción de listas, y el precio va en USD.
- Hoja 'Historial' (solo lectura): todo lo ya cargado (los 5 módulos).
"""
import io
import re
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from utils import db_utils
from utils.comisiones import comision_desde_perfil

COLS = ["Estrategia", "Operación", "Fecha", "Títulos", "Precio", "Tipo de cambio"]
_CODE = {"DCA": "DCA", "Dividendos": "DIV", "Por Objetivos": "OBJ", "FIBRAs": "FIB"}
_CODE_MOD = {"DCA": "DCA", "DIV": "Dividendos", "OBJ": "Por Objetivos", "FIB": "FIBRAs"}
_HEAD_FILL = PatternFill("solid", fgColor="6C63FF")
_HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
_TITLE_FONT = Font(bold=True, size=13, color="1A1A2E")


def _modulos():
    return [
        ("DCA", db_utils.load_strategies(), db_utils.load_purchases, None),
        ("Dividendos", db_utils.load_div_strategies(), db_utils.load_div_purchases, None),
        ("Por Objetivos", db_utils.load_obj_strategies(), db_utils.load_obj_purchases,
         db_utils.load_obj_sales),
        ("FIBRAs", db_utils.load_fibra_strategies(), db_utils.load_fibra_purchases, None),
    ]


def _copy_carteras():
    """Carteras de Copy Trading del usuario, con su etiqueta (id oculto) y los
    tickers que puede operar (los del inversionista + los que ya tenga en posición)."""
    from utils.copytrading_utils import INVERSIONISTAS
    inv_by_id = {i["id"]: i for i in INVERSIONISTAS}
    out = []
    for e in db_utils.load_copy_strategies():
        inv = inv_by_id.get(e["investor_id"])
        tickers = [t for t, _ in inv["holdings"]] if inv else []
        for p in db_utils.posiciones_copy(e["id"]):
            if p["ticker"] not in tickers:
                tickers.append(p["ticker"])
        nombre = e.get("nombre") or e["investor_id"]
        out.append({"eid": e["id"], "nombre": nombre,
                    "label": f"{nombre}  —  id COPY-{e['id']}", "tickers": tickers})
    return out


def _label_estrategia(modulo, e) -> str:
    """Etiqueta legible y ÚNICA para el desplegable (con id oculto al final)."""
    tk = e.get("ticker", "")
    code = _CODE[modulo]
    if modulo == "DCA":
        det = f"{e.get('frecuencia', '')} · inicio {str(e.get('fecha_inicio', ''))[:10]}"
        base = f"DCA · {tk} · {det}"
    elif modulo == "Por Objetivos":
        base = f"Por Objetivos · {tk} · entrada {e.get('precio_entrada')} / salida {e.get('precio_salida')}"
    elif modulo == "Dividendos":
        base = f"Dividendos · {tk}"
    else:
        base = f"FIBRAs · {tk}"
    return f"{base}  —  id {code}-{e['id']}"


# ── Generar plantilla ────────────────────────────────────────────────────────
def generar_plantilla() -> bytes:
    wb = openpyxl.Workbook()
    _hoja_instrucciones(wb.active)

    ws = wb.create_sheet("Movimientos")
    ws_lst = wb.create_sheet("Listas")

    # Lista de estrategias (para el desplegable)
    labels = [_label_estrategia(m, e) for m, ests, _lc, _lv in _modulos() for e in ests]
    ws_lst["A1"] = "Estrategias"
    for i, lab in enumerate(labels, start=2):
        ws_lst.cell(row=i, column=1, value=lab)
    ws_lst.sheet_state = "hidden"

    # Encabezados de Movimientos
    for c, h in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
        cell.alignment = Alignment(horizontal="center")
    for i, w in enumerate([50, 12, 13, 9, 11, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Desplegables
    n = max(len(labels), 1)
    dv_est = DataValidation(type="list", formula1=f"Listas!$A$2:$A${n + 1}", allow_blank=True,
                            showErrorMessage=True)
    dv_est.error = "Elige una estrategia de la lista."
    dv_est.prompt = "Elige la estrategia de la lista desplegable"
    ws.add_data_validation(dv_est)
    dv_est.add("A2:A1000")

    dv_op = DataValidation(type="list", formula1='"Compra,Venta"', allow_blank=True,
                           showErrorMessage=True)
    dv_op.error = "Escribe Compra o Venta."
    ws.add_data_validation(dv_op)
    dv_op.add("B2:B1000")

    _hoja_copy(wb, ws_lst)
    _hoja_historial(wb)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _hoja_copy(wb, ws_lst):
    """Hoja aparte para Copy Trading: Cartera + Acción (desplegables) + compra/venta.
    El precio va en USD (Copy opera en dólares)."""
    ws = wb.create_sheet("Copy Trading")
    cols = ["Cartera", "Acción", "Operación", "Fecha", "Títulos", "Precio (USD)", "Tipo de cambio"]
    for c, h in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
        cell.alignment = Alignment(horizontal="center")
    for i, w in enumerate([34, 12, 12, 13, 9, 12, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    carteras = _copy_carteras()
    if not carteras:
        ws.cell(row=2, column=1, value="No tienes carteras de Copy Trading. Copia una en la app y descarga la plantilla de nuevo.")
        return

    # Listas (hoja oculta): B = carteras, C = tickers (unión de todas)
    ws_lst["B1"] = "Carteras"
    for i, c in enumerate(carteras, start=2):
        ws_lst.cell(row=i, column=2, value=c["label"])
    tickers = sorted({t for c in carteras for t in c["tickers"]})
    ws_lst["C1"] = "Tickers"
    for i, t in enumerate(tickers, start=2):
        ws_lst.cell(row=i, column=3, value=t)

    nb, ntk = max(len(carteras), 1), max(len(tickers), 1)
    dv_cart = DataValidation(type="list", formula1=f"Listas!$B$2:$B${nb + 1}", allow_blank=True,
                             showErrorMessage=True)
    dv_cart.error = "Elige una cartera de la lista."
    ws.add_data_validation(dv_cart)
    dv_cart.add("A2:A1000")

    dv_tk = DataValidation(type="list", formula1=f"Listas!$C$2:$C${ntk + 1}", allow_blank=True,
                           showErrorMessage=True)
    dv_tk.error = "Elige una acción de la lista."
    ws.add_data_validation(dv_tk)
    dv_tk.add("B2:B1000")

    dv_op = DataValidation(type="list", formula1='"Compra,Venta"', allow_blank=True,
                           showErrorMessage=True)
    ws.add_data_validation(dv_op)
    dv_op.add("C2:C1000")


def _precio_usd(precio, tc):
    """Precio en dólares: solo aplica a empresas del SIC (compradas en USD, tc != 1)."""
    try:
        if precio is not None and tc and float(tc) != 1.0:
            return round(float(precio) / float(tc), 2)
    except Exception:
        pass
    return ""  # empresas en pesos (BMV/FIBRAs): no aplica


def _hoja_historial(wb):
    ws = wb.create_sheet("Historial")
    headers = ["Módulo", "Estrategia", "Operación", "Fecha", "Títulos",
               "Precio MXN", "Tipo de cambio", "Precio USD (SIC)", "Comisión"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
    for modulo, estrategias, load_c, load_v in _modulos():
        for e in estrategias:
            nombre = e.get("ticker", "")
            for c in load_c(e["id"]):
                tc = c.get("tipo_cambio", 1.0)
                ws.append([modulo, nombre, "Compra", c.get("fecha"), c.get("titulos"),
                           c.get("precio"), tc, _precio_usd(c.get("precio"), tc),
                           c.get("comision", 0.0)])
            if load_v:
                for v in load_v(e["id"]):
                    tc = v.get("tipo_cambio", 1.0)
                    ws.append([modulo, nombre, "Venta", v.get("fecha"), v.get("titulos"),
                               v.get("precio"), tc, _precio_usd(v.get("precio"), tc),
                               v.get("comision", 0.0)])
    # Copy Trading (canasta por acción): compras del detalle + ventas cerradas
    for cart in _copy_carteras():
        eid, nombre = cart["eid"], cart["nombre"]
        for cp in db_utils.load_copy_purchases(eid):
            tc = cp.get("tipo_cambio", 1.0)
            for d in cp["detalle"]:
                ws.append(["Copy Trading", f"{nombre} · {d['ticker']}", "Compra", cp.get("fecha"),
                           d["titulos"], round(d["precio_usd"] * tc, 2), tc,
                           round(d["precio_usd"], 2), ""])
        for v in db_utils.load_ventas_cerradas("Copy Trading", eid):
            tc = v.get("tipo_cambio", 1.0) or 1.0
            ws.append(["Copy Trading", f"{nombre} · {v['ticker']}", "Venta", v.get("fecha"),
                       v.get("titulos"), v.get("precio"), tc,
                       round((v.get("precio") or 0) / tc, 2), v.get("comision", 0.0)])
    for i, w in enumerate([14, 16, 11, 14, 10, 12, 14, 15, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _hoja_instrucciones(ws):
    ws.title = "Instrucciones"
    lineas = [
        ("Plantilla — Smart Strategy Engine", _TITLE_FONT),
        ("", None),
        ("Cómo llenarla:", Font(bold=True, size=11)),
        ("1. Ve a la pestaña 'Movimientos'.", None),
        ("2. En cada fila, elige la Estrategia y la Operación desde las listas", None),
        ("   desplegables (la flechita ▾). No las escribas a mano.", None),
        ("3. Llena los datos de cada fila:", None),
        ("     • Fecha: formato AÑO-MES-DÍA, por ejemplo 2026-01-15.", None),
        ("     • Títulos: cuántas acciones (número entero).", None),
        ("     • Precio: el precio por acción EN PESOS (MXN).", None),
        ("     • Tipo de cambio: solo si compraste en dólares (pesos por dólar). Si no, deja 1.", None),
        ("     La comisión NO se pide aquí: se calcula sola con el % de comisión de tu Perfil.", None),
        ("4. Guarda el archivo y súbelo en la app, en el botón 'Importar'.", None),
        ("", None),
        ("Notas:", Font(bold=True, size=11)),
        ("• La lista de Estrategia distingue las repetidas: si tienes dos DCA de MU,", None),
        ("  verás su frecuencia y fecha de inicio para saber cuál es cuál.", None),
        ("• En 'Movimientos' solo 'Por Objetivos' admite Venta (DCA/Dividendos/FIBRAs solo Compra).", None),
        ("", None),
        ("Copy Trading (pestaña aparte):", Font(bold=True, size=11)),
        ("• Como es una canasta, va en su propia hoja 'Copy Trading'.", None),
        ("• Elige la Cartera y la Acción de las listas, y la Operación (Compra o Venta).", None),
        ("• Aquí el Precio va EN DÓLARES (USD) por acción, porque Copy opera en dólares.", None),
        ("• El Tipo de cambio son pesos por dólar el día de tu operación.", None),
        ("", None),
        ("• La pestaña 'Historial' es solo informativa (lo que ya tienes cargado).", None),
    ]
    for i, (txt, font) in enumerate(lineas, start=1):
        ws.cell(row=i, column=1, value=txt)
        if font:
            ws.cell(row=i, column=1).font = font
    ws.column_dimensions["A"].width = 95


# ── Importar plantilla ───────────────────────────────────────────────────────
def _parse_fecha(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"fecha inválida: {v}")


def importar_excel(archivo) -> dict:
    wb = openpyxl.load_workbook(archivo, data_only=True)
    res = {"insertadas": 0, "errores": [], "por_estrategia": {}}
    if "Movimientos" not in wb.sheetnames:
        res["errores"].append("El archivo no tiene la hoja 'Movimientos'. Descarga la plantilla de nuevo.")
        return res
    ws = wb["Movimientos"]
    for r in range(2, ws.max_row + 1):
        est = ws.cell(r, 1).value
        op = ws.cell(r, 2).value
        fecha = ws.cell(r, 3).value
        tit = ws.cell(r, 4).value
        precio = ws.cell(r, 5).value
        tc = ws.cell(r, 6).value
        if not est and tit in (None, "") and precio in (None, ""):
            continue  # fila vacía
        m = re.search(r"id (DCA|DIV|OBJ|FIB)-(\d+)", str(est or ""))
        if not m:
            res["errores"].append(f"fila {r}: elige la Estrategia de la lista desplegable")
            continue
        modulo = _CODE_MOD[m.group(1)]
        sid = int(m.group(2))
        etiqueta = str(est).split("  —  ")[0]
        try:
            fecha_d = _parse_fecha(fecha)
            tit_i = int(float(tit))
            precio_f = float(precio)
            tc_f = float(tc) if tc not in (None, "") else 1.0
            # La comisión ya no se pide en el Excel: se calcula con el % del perfil.
            com_f = comision_desde_perfil(tit_i * precio_f)
            es_venta = bool(op) and str(op).strip().lower().startswith("v")
        except Exception:
            res["errores"].append(f"fila {r}: datos incompletos o inválidos")
            continue
        if tit_i <= 0 or precio_f <= 0:
            res["errores"].append(f"fila {r}: títulos y precio deben ser mayores a 0")
            continue
        ok = _insertar(modulo, sid, es_venta, fecha_d, tit_i, precio_f, tc_f, com_f, res, r)
        if ok:
            res["insertadas"] += 1
            res["por_estrategia"][etiqueta] = res["por_estrategia"].get(etiqueta, 0) + 1

    if "Copy Trading" in wb.sheetnames:
        _importar_copy(wb["Copy Trading"], res)
    return res


def _importar_copy(ws, res):
    """Importa la hoja 'Copy Trading': Cartera + Acción + Compra/Venta (precio en USD)."""
    for r in range(2, ws.max_row + 1):
        cart = ws.cell(r, 1).value
        tk = ws.cell(r, 2).value
        op = ws.cell(r, 3).value
        fecha = ws.cell(r, 4).value
        tit = ws.cell(r, 5).value
        precio = ws.cell(r, 6).value  # USD
        tc = ws.cell(r, 7).value
        if not cart and tit in (None, "") and precio in (None, ""):
            continue  # fila vacía
        m = re.search(r"id COPY-(\d+)", str(cart or ""))
        if not m:
            res["errores"].append(f"Copy fila {r}: elige la cartera de la lista desplegable")
            continue
        eid = int(m.group(1))
        if not tk:
            res["errores"].append(f"Copy fila {r}: elige la acción de la lista")
            continue
        ticker = str(tk).strip().upper()
        try:
            fecha_d = _parse_fecha(fecha)
            tit_i = int(float(tit))
            precio_usd = float(precio)
            tc_f = float(tc) if tc not in (None, "") else 1.0
            es_venta = bool(op) and str(op).strip().lower().startswith("v")
        except Exception:
            res["errores"].append(f"Copy fila {r}: datos incompletos o inválidos")
            continue
        if tit_i <= 0 or precio_usd <= 0:
            res["errores"].append(f"Copy fila {r}: títulos y precio deben ser mayores a 0")
            continue
        com = comision_desde_perfil(tit_i * precio_usd * tc_f)
        etiqueta = f"{str(cart).split('  —  ')[0]} · {ticker}"
        if es_venta:
            r2 = db_utils.registrar_venta_copy(eid, ticker, fecha_d, tit_i, precio_usd, tc_f, com)
            if not r2.get("ok"):
                res["errores"].append(f"Copy fila {r}: {r2.get('msg')}")
                continue
        else:
            db_utils.save_copy_purchase(eid, fecha_d, tit_i * precio_usd * tc_f, tc_f,
                                        [{"ticker": ticker, "titulos": tit_i, "precio_usd": precio_usd}],
                                        comision=com)
        res["insertadas"] += 1
        res["por_estrategia"][etiqueta] = res["por_estrategia"].get(etiqueta, 0) + 1


def _insertar(modulo, sid, es_venta, fecha, tit, precio, tc, com, res, r) -> bool:
    if modulo == "DCA":
        if es_venta:
            res["errores"].append(f"fila {r}: DCA no admite ventas")
            return False
        db_utils.save_purchase(sid, fecha, tit, precio, tc, com)
        return True
    if modulo == "Dividendos":
        if es_venta:
            res["errores"].append(f"fila {r}: Dividendos no admite ventas")
            return False
        db_utils.save_div_purchase(sid, fecha, tit, precio, tc, com)
        return True
    if modulo == "FIBRAs":
        if es_venta:
            res["errores"].append(f"fila {r}: FIBRAs no admite ventas")
            return False
        db_utils.save_fibra_purchase(sid, fecha, tit, precio, com)
        return True
    if modulo == "Por Objetivos":
        if not es_venta:
            db_utils.save_obj_purchase(sid, fecha, tit, precio, tc, com)
            return True
        return _vender_objetivos_fifo(sid, fecha, tit, precio, tc, com, res, r)
    res["errores"].append(f"fila {r}: módulo desconocido")
    return False


def _vender_objetivos_fifo(sid, fecha, titulos, precio, tc, com, res, r) -> bool:
    compras = sorted(db_utils.load_obj_purchases(sid), key=lambda c: str(c["fecha"]))
    ventas = db_utils.load_obj_sales(sid)
    vendido = {}
    for v in ventas:
        vendido[v["compra_id"]] = vendido.get(v["compra_id"], 0) + v["titulos"]
    restante = titulos
    com_pendiente = com
    for c in compras:
        if restante <= 0:
            break
        disponible = c["titulos"] - vendido.get(c["id"], 0)
        if disponible <= 0:
            continue
        asignar = min(disponible, restante)
        db_utils.save_obj_sale(c["id"], sid, fecha, asignar, precio, tc, com_pendiente)
        com_pendiente = 0.0
        restante -= asignar
    if restante == titulos:
        res["errores"].append(f"fila {r}: no hay títulos disponibles para vender")
        return False
    if restante > 0:
        res["errores"].append(
            f"fila {r}: venta de {titulos} excede lo disponible ({titulos - restante} aplicados)")
    return True
